from django.contrib import admin
from django.http import HttpResponse
from invoices.models import Invoice

import datetime

date = datetime.datetime.now().strftime("%Y-%m-%d")

def export_csv(modeladmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(mimetype='text/csv')
    response['Content-Disposition'] = 'attachment; filename=3DP-' + date +'.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"Filename"),
        smart_str(u"Created"),
        smart_str(u"Updated"),
        smart_str(u"Customer First Name"),
        smart_str(u"Customer Last Name"),
        smart_str(u"Customer Email"),
        smart_str(u"Customer Phone"),
        smart_str(u"Customer Status"),
        smart_str(u"Customer Filename"),
        smart_str(u"Customer Purpose"),
        smart_str(u"Customer Department"),
        smart_str(u"SD Card Number"),
        smart_str(u"Estimated Time"),
        smart_str(u"Estimated Cost"),
        smart_str(u"Actual Time"),
        smart_str(u"Job Cost"),
        smart_str(u"Description"),
        smart_str(u"Job State"),
    ])

    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.filename),
            smart_str(obj.created),
            smart_str(obj.updated),
            smart_str(obj.customer_first_name),
            smart_str(obj.customer_last_name),
            smart_str(obj.customer_email),
            smart_str(obj.customer_phone),
            smart_str(obj.customer_status),
            smart_str(obj.customer_filename),
            smart_str(obj.customer_purpose),
            smart_str(obj.customer_department),
            smart_str(obj.sd_card_number),
            smart_str(obj.estimated_time),
            smart_str(obj.estimated_cost),
            smart_str(obj.actual_time),
            smart_str(obj.job_cost),
            smart_str(obj.description),
            smart_str(obj.job_state),
        ])
    return response
export_csv.short_description = u"Export CSV"

def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from openpyxl.cell import get_column_letter
    response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=3DP-' + date +'.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "3DP " + date

    row_num = 1

    columns = [
        (u"ID", 4),
        (u"Filename", 20),
        (u"Created", 10),
        (u"Updated", 10),
        (u"Customer First Name", 20),
        (u"Customer Last Name", 20),
        (u"Customer Email", 30),
        (u"Customer Phone", 15),
        (u"Customer Status", 15),
        (u"Customer Filename", 30),
        (u"Customer Purpose", 20),
        (u"Customer Dept", 15),
        (u"SD Card Number", 15),
        (u"Estimated Time", 15),
        (u"Estimated Cost", 15),
        (u"Actual Time", 15),
        (u"Job Cost", 15),
        (u"Description", 70),
        (u"Job State", 10),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num, column=col_num+1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.filename,
            obj.created,
            obj.updated,
            obj.customer_first_name,
            obj.customer_last_name,
            obj.customer_email,
            obj.customer_phone,
            obj.customer_status,
            obj.customer_filename,
            obj.customer_purpose,
            obj.customer_department,
            obj.sd_card_number,
            obj.estimated_time,
            obj.estimated_cost,
            obj.actual_time,
            obj.job_cost,
            obj.description,
            obj.job_state,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num, column=col_num+1)
            c.value = row[col_num]
            #c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"

class InvoiceAdmin(admin.ModelAdmin):
    list_display = ('customer_filename', 'created', 'job_state')
    actions = [export_xlsx, export_csv]


admin.site.register(Invoice, InvoiceAdmin)